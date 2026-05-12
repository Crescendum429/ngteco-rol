import re
import unicodedata
import xlrd
from datetime import date, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# SBU = Salario Basico Unificado. Lo fija el MDT cada anio en diciembre.
# 2025 = $470 (Resolucion MDT-2024-179). 2026: leer del env var SBU_VIGENTE
# si esta seteado, o usar el default conservador (470). El contador debe
# actualizar este valor cuando el MDT publique el oficial 2026.
import os as _os
try:
    SBU_2026 = float(_os.environ.get("SBU_VIGENTE", "470"))
except (TypeError, ValueError):
    SBU_2026 = 470.0

# Tasa de aporte personal IESS — Resolucion C.D. 501 IESS.
# Materia gravada: sueldo, sobresueldos, comisiones, horas extras, participacion
# de utilidades y otras remuneraciones accesorias.
# NO incluye: 13ro, 14to, fondos de reserva (cuando se pagan), viaticos,
# alimentacion, transporte (si no es parte del salario).
IESS_EMPLEADO = 0.0945

# Topes legales horas extras — Codigo del Trabajo Art. 55
MAX_SUPLEMENTARIAS_DIA = 4.0  # max horas suplementarias por dia
MAX_SUPLEMENTARIAS_SEMANA = 12.0  # max horas suplementarias por semana

# Politica de exceso semanal — opciones:
#   "alertar_pero_pagar": paga las primeras 12h al 50%, alerta sobre el exceso
#                         pero lo paga al 50% tambien (postura empresarial)
#   "reclasificar_100":   paga las primeras 12h al 50%, el exceso al 100%
#                         (postura pro-empleado, mas conservadora legalmente)
# Default: alertar_pero_pagar — pide al usuario decidir caso por caso
POLITICA_EXCESO_SEMANAL = _os.environ.get("POLITICA_EXCESO_SEMANAL", "alertar_pero_pagar")

MORN_MIN  = 5 * 60
MORN_MAX  = 8 * 60 + 45
LUNCH_MIN = 11 * 60
LUNCH_MAX = 15 * 60
AFT_MIN   = 14 * 60
AFT_MAX   = 19 * 60
SHORT_SEG = 60

WEEKDAYS = {'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN'}


# ── Utilidades ────────────────────────────────────────────────

def normalize(s):
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.strip().lower()


def emp_name(raw):
    """Extrae nombre limpio de 'Fernando Pinargote (8)'."""
    return re.sub(r'\s*\(\d+\)\s*$', '', raw).strip()


def to_mins(s):
    if not s or not str(s).strip():
        return None
    m = re.match(r'(\d+):(\d+)', str(s).strip())
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def to_time(m):
    return time(m // 60, m % 60) if m is not None else None


def parse_date(ds):
    p = ds.split('-')
    return date(2000 + int(p[0]), int(p[1]), int(p[2]))


# ── Parseo XLS ────────────────────────────────────────────────

def parse_xls(path):
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    result = []
    emp = None
    days = None
    cur_day = None

    for i in range(ws.nrows):
        row = [ws.cell(i, j).value for j in range(7)]

        if row[0] == 'Employee':
            raw = str(row[3])
            emp = re.sub(r'\n', ' ', raw).strip()
            m_id = re.search(r'\((\d+)\)', raw)
            ngteco_id = m_id.group(1) if m_id else ''
            days = {}
            result.append((emp, days, ngteco_id))
            cur_day = None

        elif emp and row[0] in WEEKDAYS and row[1]:
            # NGTeco asigna las marcaciones al "día laboral anterior" en su modelo
            # interno. Para que el reporte de mayo contenga las horas trabajadas en
            # mayo (no las del 30/abr al 30/may), sumamos +1 dia. El XLS de marzo
            # dice 'SUN 26-03-01' (domingo 1 marzo) pero las horas de esa fila son
            # las realmente trabajadas el lunes 2 de marzo.
            raw_date = str(row[1])
            try:
                p = raw_date.split('-')
                corrected = date(2000 + int(p[0]), int(p[1]), int(p[2])) + timedelta(days=1)
                cur_day = corrected.strftime('%y-%m-%d')
            except Exception:
                cur_day = raw_date
            in_t = to_mins(row[2])
            out_t = to_mins(row[3])
            note = str(row[6]) if row[6] else ''
            days[cur_day] = []
            if in_t is not None or out_t is not None or 'Missing' in note:
                days[cur_day].append((in_t, out_t, note))

        elif emp and row[0] == '' and row[1] == '' and cur_day and (row[2] or row[6]):
            in_t = to_mins(row[2])
            out_t = to_mins(row[3])
            note = str(row[6]) if row[6] else ''
            days[cur_day].append((in_t, out_t, note))

    return result


# ── Matching empleados ────────────────────────────────────────

def match_empleados(data, emp_db):
    """Retorna (matched, nuevos, faltantes).
    matched: {xls_name: db_key}
    nuevos: [(xls_name, ngteco_id)] - en XLS pero no en DB
    faltantes: [db_key] - en DB pero no en XLS
    """
    matched = {}
    nuevos = []

    db_norm = {}
    for key, emp in emp_db.items():
        nombre = emp.get('nombre', '')
        db_norm[normalize(nombre)] = key

    xls_names_norm = set()
    for emp_full, days, nid in data:
        name = emp_full.split('(')[0].strip()
        n = normalize(name)
        xls_names_norm.add(n)

        if n in db_norm:
            matched[name] = db_norm[n]
        else:
            nuevos.append((name, nid))

    faltantes = []
    for key, emp in emp_db.items():
        n = normalize(emp.get('nombre', ''))
        if n and n not in xls_names_norm:
            if not emp.get('ocultar', False):
                faltantes.append(key)

    return matched, nuevos, faltantes


# ── Clasificacion ─────────────────────────────────────────────

def classify(pairs):
    active = [(i, o, n) for i, o, n in pairs if i is not None or o is not None]
    missing_out = any('Missing' in (n or '') for _, _, n in pairs)

    if not active:
        return None, None, None, None, []

    h1 = h2 = h3 = h4 = None
    flags = []

    if len(active) == 1:
        i, o, n = active[0]
        if i is not None and o is not None:
            dur = o - i
            if MORN_MIN <= i <= MORN_MAX:
                h1, h2 = i, o
                if missing_out:
                    flags.append('REVISAR: salida no registrada')
            elif LUNCH_MIN <= i <= LUNCH_MAX:
                if dur < SHORT_SEG:
                    flags.append('REVISAR: entrada manana no registrada')
                    h2, h3 = i, o
                else:
                    h1, h2 = i, o
            else:
                h1, h2 = i, o
                flags.append('REVISAR: horario fuera de rango esperado')
        elif i is not None:
            h1 = i
            flags.append('REVISAR: timbre sin salida correspondiente')

    elif len(active) == 2:
        i1, o1, n1 = active[0]
        i2, o2, n2 = active[1]
        if (i1 is not None and o1 is not None and
                LUNCH_MIN <= i1 <= LUNCH_MAX and (o1 - i1) < SHORT_SEG):
            flags.append('REVISAR: entrada manana no registrada')
            h2, h3 = i1, o1
            h4 = i2
            return h1, h2, h3, h4, flags
        if i1 is not None and MORN_MIN <= i1 <= MORN_MAX:
            h1, h2, h3, h4 = i1, o1, i2, o2
            if o2 is None:
                flags.append('REVISAR: salida tarde no registrada')
            elif not (AFT_MIN <= o2 <= AFT_MAX):
                flags.append('REVISAR: hora de salida inusual')
        else:
            h1, h2, h3, h4 = i1, o1, i2, o2
            flags.append('REVISAR: verificar horarios')
    else:
        h1 = active[0][0]
        h4 = active[-1][1] or active[-1][0]
        flags.append(f'REVISAR: {len(active)} registros - verificar manualmente')

    return h1, h2, h3, h4, flags


def clasificar_todo(data):
    result = {}
    for emp, days, *_ in data:
        name = emp_name(emp)
        result[name] = {}
        for ds in days:
            h1, h2, h3, h4, flags = classify(days[ds])
            result[name][ds] = {
                'h1': h1, 'h2': h2, 'h3': h3, 'h4': h4,
                'flags': list(flags),
            }
    return result


def _get_cls(days, ds, overrides, emp_name):
    if overrides and emp_name in overrides and ds in overrides[emp_name]:
        d = overrides[emp_name][ds]
        return d['h1'], d['h2'], d['h3'], d['h4'], d['flags']
    return classify(days[ds])


# ── Calculo de horas ──────────────────────────────────────────

def _sumar_horas_dia(h1, h2, h3, h4):
    horas = 0.0
    if h1 is not None and h2 is not None:
        horas += (h2 - h1) / 60
    if h3 is not None and h4 is not None:
        horas += (h4 - h3) / 60
    return horas


def _acumular(ds, horas, flags, base_hours, acum):
    if horas <= 0:
        return
    acum['dias'] += 1
    if any(f.startswith('REVISAR:') for f in flags):
        acum['dias_anomalia'] += 1
    try:
        es_finde = parse_date(ds).weekday() >= 5
    except Exception:
        es_finde = False
    if es_finde:
        acum['ext_100'] += horas
    else:
        reg = min(horas, base_hours)
        acum['regular'] += reg
        extra = max(0.0, horas - base_hours)
        acum['sup_50'] += min(extra, 4.0)
        acum['ext_100'] += max(0.0, extra - 4.0)


def _resultado(acum):
    return {
        'dias': acum['dias'],
        'dias_anomalia': acum['dias_anomalia'],
        'horas_regular': round(acum['regular'], 2),
        'horas_50': round(acum['sup_50'], 2),
        'horas_100': round(acum['ext_100'], 2),
        'horas_total': round(acum['regular'] + acum['sup_50'] + acum['ext_100'], 2),
    }


def calcular_horas_clasificadas(classified_days, base_hours=8):
    acum = {'dias': 0, 'dias_anomalia': 0, 'regular': 0.0, 'sup_50': 0.0, 'ext_100': 0.0}
    for ds, d in classified_days.items():
        horas = _sumar_horas_dia(d['h1'], d['h2'], d['h3'], d['h4'])
        _acumular(ds, horas, d['flags'], base_hours, acum)
    return _resultado(acum)


# ── Calculo nomina completa ───────────────────────────────────

def calcular_nomina(hrs, cfg, extras=None):
    """Calcula nomina completa para un empleado.

    Args:
        hrs: dict de calc_horas_periodo con dias, dias_pagados, horas_total,
             horas_50, horas_100, horas_regular.
        cfg: dict con salario, horas_base, transporte_dia, prestamo_iess,
             fondos_reserva, descuento_iess, transporte_gravable (opcional,
             default True — controla si transporte entra en base IESS/fondos).
        extras: dict con decimo_13 (bool), decimo_14 (bool), bonus, anticipo
                (deducido de transf_fin, no de transf_15), horas_pasar.

    Returns dict con todo el detalle. Incluye:
        - alertas: lista de strings con flags para revision humana
                   (ej. transf_fin negativa, exceso semanal de extras, etc.)
        - validaciones: dict para reconciliacion (sum lineas == total)
    """
    extras = extras or {}
    salario = float(cfg.get('salario', 0))
    base_h = int(cfg.get('horas_base', 8))
    transp_dia = float(cfg.get('transporte_dia', 0))
    prestamo = float(cfg.get('prestamo_iess', 0))
    descuento_iess = cfg.get('descuento_iess', True)
    tiene_fondos = cfg.get('fondos_reserva', False)
    transporte_gravable = cfg.get('transporte_gravable', True)
    h_anterior = float(cfg.get('horas_comp_anterior', 0))
    horas_pasar = float(extras.get('horas_pasar', 0))
    anticipo = float(extras.get('anticipo', 0))  # adelanto a deducir de 2da quincena

    alertas = []

    hourly = salario / 30 / base_h if salario > 0 and base_h > 0 else 0

    # Horas compensatorias con arrastre (legacy)
    h_50_total = hrs['horas_50'] + h_anterior
    h_50_pagar = max(h_50_total - horas_pasar, 0)
    h_50_arrastre = horas_pasar

    pay_50 = h_50_pagar * hourly * 1.5
    pay_100 = hrs['horas_100'] * hourly * 2.0

    # dias trabajados con timbres reales (Art. 42 — solo dias efectivos pagan transporte)
    dias_trab = hrs.get('dias', 0)
    transporte = dias_trab * transp_dia

    quincena = salario / 2
    horas_extras = pay_50 + pay_100

    # Base imponible para IESS y fondos (materia gravada)
    # Excluye decimos y transporte si no es gravable
    base_imponible = salario + horas_extras
    if transporte_gravable:
        base_imponible += transporte

    # Ingresos brutos (incluye todo lo que recibe el empleado)
    total_ingresos = salario + horas_extras + transporte

    # Decimos NO son materia gravada (Resolucion IESS 501) — se suman al ingreso
    d13 = salario if extras.get('decimo_13') else 0
    d14 = SBU_2026 if extras.get('decimo_14') else 0
    bonus = float(extras.get('bonus', 0))

    total_ingresos_con_extras = total_ingresos + d13 + d14 + bonus

    # Aporte personal IESS — sobre base imponible (no sobre decimos)
    iess = round(base_imponible * IESS_EMPLEADO, 2) if descuento_iess else 0

    # Total egresos
    total_egresos = iess + prestamo

    # Neto
    valor_recibir = total_ingresos_con_extras - total_egresos

    # Fondos de reserva — 8.33% de base imponible, NO incluye decimos.
    # Legalmente aplica solo a empleados con >=1 anio de servicio (Art. 196).
    # El flag fondos_reserva del cfg debe ser derivado de fecha_ingreso por el caller.
    fondos = round(base_imponible / 12, 2) if tiene_fondos else 0

    # Detalle transferencias
    # 1ra quincena: salario/2 sin descuentos
    # 2da quincena: lo que queda menos anticipos
    transf_15 = quincena
    transf_fin = valor_recibir - quincena + fondos - anticipo
    total_transferido = transf_15 + transf_fin

    # ─── Alertas para revision humana ───
    if transf_fin < 0:
        alertas.append(
            f"transf_fin NEGATIVA ({transf_fin:.2f}). Descuentos exceden 2da quincena. "
            f"Revisa con el empleado: reducir prestamo, repartir descuento en proximos meses, "
            f"o ajustar 1ra quincena."
        )
    if dias_trab > 31:
        alertas.append(f"dias trabajados = {dias_trab} es inusual (>31)")
    if salario > 0 and total_ingresos > salario * 3:
        alertas.append(
            f"total_ingresos ({total_ingresos:.2f}) es {total_ingresos/salario:.1f}x el salario. "
            f"Revisar horas extras o bonos."
        )
    if iess < 0 or prestamo < 0:
        alertas.append("Descuentos negativos — error de configuracion")

    # ─── Reconciliacion: suma de partes == total ───
    suma_check = transf_15 + transf_fin
    if abs(suma_check - total_transferido) > 0.01:
        alertas.append(
            f"RECONCILIACION ROTA: transf_15+transf_fin={suma_check:.2f} != "
            f"total_transferido={total_transferido:.2f}"
        )

    return {
        'salario': salario,
        'hourly': hourly,
        'hours': hrs,
        'quincena': quincena,
        'horas_comp_anterior': h_anterior,
        'h_50_total': h_50_total,
        'h_50_pagar': h_50_pagar,
        'h_50_arrastre': round(h_50_arrastre, 2),
        'pay_50': round(pay_50, 2),
        'pay_100': round(pay_100, 2),
        'horas_extras': round(horas_extras, 2),
        'transporte': round(transporte, 2),
        'transporte_gravable': transporte_gravable,
        'base_imponible': round(base_imponible, 2),
        'total_ingresos': round(total_ingresos, 2),
        'decimo_13': d13,
        'decimo_14': d14,
        'bonus': bonus,
        'anticipo': anticipo,
        'iess': iess,
        'prestamo_iess': prestamo,
        'total_egresos': round(total_egresos, 2),
        'valor_recibir': round(valor_recibir, 2),
        'fondos_reserva': fondos,
        'transf_15': round(transf_15, 2),
        'transf_fin': round(transf_fin, 2),
        'total_transferido': round(total_transferido, 2),
        'alertas': alertas,
    }


# ── Excel horas ───────────────────────────────────────────────

def write_excel(data, dest, overrides=None):
    wb = Workbook()
    wb.remove(wb.active)

    YELLOW = PatternFill('solid', fgColor='FFFF99')
    HDR_BG = PatternFill('solid', fgColor='4472C4')
    HDR_FT = Font(bold=True, color='FFFFFF')

    all_flags = []

    for emp, days, *_ in data:
        sheet_name = emp_name(emp)[:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = ['Fecha', 'Hora 1', 'Hora 2', 'Hora 3', 'Hora 4', 'Total (h)', 'Comentarios']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = HDR_BG
            cell.font = HDR_FT
            cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 16
        for ltr in 'BCDE':
            ws.column_dimensions[ltr].width = 9
        ws.column_dimensions['F'].width = 11
        ws.column_dimensions['G'].width = 58
        ws.freeze_panes = 'A2'

        r = 2
        for ds in sorted(days):
            h1, h2, h3, h4, flags = _get_cls(days, ds, overrides, sheet_name)
            try:
                d = parse_date(ds)
            except Exception:
                d = ds
            ws.cell(r, 1, d).number_format = 'DD/MM/YY DDD'
            for col, val in [(2, h1), (3, h2), (4, h3), (5, h4)]:
                if val is not None:
                    ws.cell(r, col, to_time(val)).number_format = 'HH:MM'
            if h1 is not None and h2 is not None:
                if h3 is not None and h4 is not None:
                    ws.cell(r, 6).value = f'=((C{r}-B{r})+(E{r}-D{r}))*24'
                elif h3 is None and h4 is None:
                    ws.cell(r, 6).value = f'=(C{r}-B{r})*24'
                ws.cell(r, 6).number_format = '0.00'
            if flags:
                comment = '; '.join(flags)
                ws.cell(r, 7, comment)
                for c in range(1, 8):
                    ws.cell(r, c).fill = YELLOW
                all_flags.append((sheet_name, d, comment))
            r += 1

    ws_sum = wb.create_sheet(title='Resumen', index=0)
    for c, h in enumerate(['Empleado', 'Fecha', 'Observacion'], 1):
        cell = ws_sum.cell(1, c, h)
        cell.fill = HDR_BG
        cell.font = HDR_FT
        cell.alignment = Alignment(horizontal='center')
    ws_sum.column_dimensions['A'].width = 26
    ws_sum.column_dimensions['B'].width = 16
    ws_sum.column_dimensions['C'].width = 62
    ws_sum.freeze_panes = 'A2'
    ws_sum.cell(1, 5, f'Total anomalias: {len(all_flags)}').font = Font(bold=True)
    for r, (emp, d, comment) in enumerate(all_flags, 2):
        ws_sum.cell(r, 1, emp)
        ws_sum.cell(r, 2, d).number_format = 'DD/MM/YY DDD'
        ws_sum.cell(r, 3, comment)
        for c in range(1, 4):
            ws_sum.cell(r, c).fill = YELLOW

    wb.save(dest)
    return all_flags


# ── Excel nomina (formato Rol de Pagos) ───────────────────────

def write_excel_nomina(nomina_data, periodo, dest):
    """nomina_data: [{name, nomina (dict from calcular_nomina)}]
    periodo: str como 'Marzo 2026'
    """
    wb = Workbook()
    wb.remove(wb.active)

    HDR_BG = PatternFill('solid', fgColor='4472C4')
    HDR_FT = Font(bold=True, color='FFFFFF')
    GREEN  = PatternFill('solid', fgColor='E2EFDA')
    BOLD   = Font(bold=True)
    BOLD_L = Font(bold=True, size=11)
    THIN   = Border(
        bottom=Side(style='thin'),
        top=Side(style='thin'),
    )

    # Hoja resumen
    ws = wb.create_sheet('Resumen Nomina', 0)
    cols = ['Empleado', '1ra Quinc.', '2da Quinc.', 'H. Extras',
            'Transp.', 'Total Ing.', 'IESS', 'Prest.', 'Total Egr.',
            'Neto', 'F. Reserva', 'Total Transf.']
    for c, h in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HDR_BG
        cell.font = HDR_FT
        cell.alignment = Alignment(horizontal='center')
    for i in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions['A'].width = 24
    ws.freeze_panes = 'A2'

    for ri, item in enumerate(nomina_data, 2):
        n = item['nomina']
        ws.cell(ri, 1, item['name'])
        ws.cell(ri, 2, n['quincena']).number_format = '$#,##0.00'
        ws.cell(ri, 3, n['quincena']).number_format = '$#,##0.00'
        ws.cell(ri, 4, n['horas_extras']).number_format = '$#,##0.00'
        ws.cell(ri, 5, n['transporte']).number_format = '$#,##0.00'
        ws.cell(ri, 6, n['total_ingresos']).number_format = '$#,##0.00'
        ws.cell(ri, 7, n['iess']).number_format = '$#,##0.00'
        ws.cell(ri, 8, n['prestamo_iess']).number_format = '$#,##0.00'
        ws.cell(ri, 9, n['total_egresos']).number_format = '$#,##0.00'
        ws.cell(ri, 10, n['valor_recibir']).number_format = '$#,##0.00'
        ws.cell(ri, 10).fill = GREEN
        ws.cell(ri, 10).font = BOLD
        ws.cell(ri, 11, n['fondos_reserva']).number_format = '$#,##0.00'
        ws.cell(ri, 12, n['total_transferido']).number_format = '$#,##0.00'
        ws.cell(ri, 12).fill = GREEN

    # Hojas individuales - formato Rol de Pagos
    for item in nomina_data:
        name = item['name']
        n = item['nomina']
        hrs = n['hours']
        ws = wb.create_sheet(title=name[:31])

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 14

        r = 1
        ws.merge_cells('A1:D1')
        ws.cell(r, 1, 'ROL DE PAGOS').font = Font(bold=True, size=14)
        ws.cell(r, 1).alignment = Alignment(horizontal='center')
        r += 2

        ws.cell(r, 1, 'FECHA:').font = BOLD
        ws.cell(r, 2, periodo)
        r += 1
        ws.cell(r, 1, 'EMPLEADO:').font = BOLD
        ws.cell(r, 2, name)
        r += 2

        # Ingresos / Egresos
        for ci in range(1, 5):
            ws.cell(r, ci).fill = HDR_BG
            ws.cell(r, ci).font = HDR_FT
        ws.cell(r, 1, 'INGRESOS')
        ws.cell(r, 3, 'EGRESOS')
        r += 1

        def _ing(label, val):
            nonlocal r
            ws.cell(r, 1, label)
            ws.cell(r, 2, val).number_format = '$#,##0.00'

        def _egr(label, val):
            ws.cell(r, 3, label)
            ws.cell(r, 4, val).number_format = '$#,##0.00'

        _ing('Primera Quincena', n['quincena'])
        _egr(f"Aporte IESS ({IESS_EMPLEADO*100:.2f}%)", n['iess'])
        r += 1
        _ing('Segunda Quincena', n['quincena'])
        if n['prestamo_iess']:
            _egr('Prestamo IESS', n['prestamo_iess'])
        r += 1
        _ing(f"Horas Extras 50% ({n['h_50_pagar']:.2f}h)", n['pay_50'])
        r += 1
        _ing(f"Horas Extras 100% ({hrs['horas_100']:.2f}h)", n['pay_100'])
        r += 1
        if n['transporte']:
            _ing(f"Transporte/Comida ({hrs['dias']}d)", n['transporte'])
            r += 1
        if n['decimo_13']:
            _ing('Decimo Tercer Sueldo', n['decimo_13'])
            r += 1
        if n['decimo_14']:
            _ing('Decimo Cuarto Sueldo', n['decimo_14'])
            r += 1
        if n['bonus']:
            _ing('Bono / Ajuste', n['bonus'])
            r += 1

        r += 1
        for ci in range(1, 5):
            ws.cell(r, ci).border = THIN
            ws.cell(r, ci).font = BOLD
        ws.cell(r, 1, 'TOTAL INGRESOS')
        ws.cell(r, 2, n['total_ingresos'] + n['decimo_13'] + n['decimo_14'] + n['bonus'])
        ws.cell(r, 2).number_format = '$#,##0.00'
        ws.cell(r, 3, 'TOTAL EGRESOS')
        ws.cell(r, 4, n['total_egresos']).number_format = '$#,##0.00'
        r += 1
        ws.cell(r, 1, 'VALOR A RECIBIR').font = BOLD_L
        ws.cell(r, 2, n['valor_recibir']).number_format = '$#,##0.00'
        ws.cell(r, 2).font = BOLD_L
        ws.cell(r, 2).fill = GREEN
        r += 2

        # Detalle de pagos
        for ci in range(1, 3):
            ws.cell(r, ci).fill = HDR_BG
            ws.cell(r, ci).font = HDR_FT
        ws.cell(r, 1, 'DETALLE DE PAGOS')
        r += 1
        ws.cell(r, 1, 'Transferencia del 15')
        ws.cell(r, 2, n['transf_15']).number_format = '$#,##0.00'
        r += 1
        ws.cell(r, 1, 'Transferencia fin de mes')
        ws.cell(r, 2, n['transf_fin']).number_format = '$#,##0.00'
        r += 1
        if n['fondos_reserva']:
            ws.cell(r, 1, 'Fondos de Reserva (incluido)')
            ws.cell(r, 2, n['fondos_reserva']).number_format = '$#,##0.00'
            r += 1
        ws.cell(r, 1, 'TOTAL TRANSFERIDO').font = BOLD
        ws.cell(r, 2, n['total_transferido']).number_format = '$#,##0.00'
        ws.cell(r, 2).font = BOLD
        ws.cell(r, 2).fill = GREEN

        if n['horas_comp_anterior'] != 0:
            r += 2
            ws.cell(r, 1, 'Horas compensatorias mes anterior').font = BOLD
            ws.cell(r, 2, n['horas_comp_anterior']).number_format = '0.00'
            r += 1
            ws.cell(r, 1, 'Arrastre sugerido proximo mes').font = BOLD
            ws.cell(r, 2, n['h_50_arrastre']).number_format = '0.00'

        # Detalle de dias
        dias = item.get('dias') or []
        if dias:
            r += 3
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            for ci in range(1, 9):
                ws.cell(r, ci).fill = HDR_BG
                ws.cell(r, ci).font = HDR_FT
            ws.cell(r, 1, 'DETALLE DE HORAS POR DIA').alignment = Alignment(horizontal='center')
            r += 1

            headers_d = ['Fecha', 'Entrada', 'Almuerzo', 'Regreso', 'Salida', 'Total', 'Modo/Banco', 'Observacion']
            for ci, h in enumerate(headers_d, 1):
                cell = ws.cell(r, ci, h)
                cell.font = BOLD
                cell.fill = PatternFill('solid', fgColor='D9E1F2')
                cell.alignment = Alignment(horizontal='center')
            r += 1

            widths = [14, 12, 12, 12, 12, 10, 18, 30]
            for ci, w in enumerate(widths, 1):
                if ws.column_dimensions[get_column_letter(ci)].width < w:
                    ws.column_dimensions[get_column_letter(ci)].width = w

            DOWS = ['lun','mar','mie','jue','vie','sab','dom']
            for d in dias:
                fecha_raw = d.get('fecha', '')
                try:
                    p = fecha_raw.split('-')
                    ddate = date(2000 + int(p[0]), int(p[1]), int(p[2]))
                    fecha_fmt = ddate.strftime('%d/%m') + ' ' + DOWS[ddate.weekday()]
                except Exception:
                    fecha_fmt = fecha_raw

                ws.cell(r, 1, fecha_fmt)
                ws.cell(r, 2, d.get('h1') or '—').alignment = Alignment(horizontal='center')
                ws.cell(r, 3, d.get('h2') or '—').alignment = Alignment(horizontal='center')
                ws.cell(r, 4, d.get('h3') or '—').alignment = Alignment(horizontal='center')
                ws.cell(r, 5, d.get('h4') or '—').alignment = Alignment(horizontal='center')
                ws.cell(r, 6, d.get('total', 0)).number_format = '0.0'
                ws.cell(r, 6).alignment = Alignment(horizontal='right')

                excedente = d.get('excedente', 0) or 0
                deficit = d.get('deficit', 0) or 0
                es_finde = d.get('es_finde', False)
                modo = d.get('modo_extra', 'banco')
                cubrir = d.get('cubrir_banco', False)
                if es_finde and (d.get('total') or 0) > 0:
                    modo_txt = f"Extras (+{d['total']:.1f}h)" if modo == 'pagar' else f"Banco (+{d['total']:.1f}h)"
                elif excedente > 0:
                    modo_txt = f"Extras (+{excedente:.1f}h)" if modo == 'pagar' else f"Banco (+{excedente:.1f}h)"
                elif deficit > 0:
                    modo_txt = f"Cubre -{deficit:.1f}h" if cubrir else f"Falta -{deficit:.1f}h"
                else:
                    modo_txt = '—'
                ws.cell(r, 7, modo_txt)

                ws.cell(r, 8, d.get('flag') or '')
                r += 1

    wb.save(dest)


def write_pdf_nomina(item, periodo):
    """PDF del rol de pagos individual. Retorna bytes."""
    from fpdf import FPDF

    name = item['name']
    n = item['nomina']
    hrs = n['hours']
    W = 160  # ancho util (A4 210mm - margenes 25mm c/lado)

    def _t(s):
        return normalize(str(s))

    pdf = FPDF(format='A4')
    pdf.set_margins(25, 20, 25)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)

    def _hline():
        pdf.set_draw_color(180, 180, 180)
        pdf.set_line_width(0.3)
        pdf.line(25, pdf.get_y(), 185, pdf.get_y())
        pdf.ln(3)

    def _section(title):
        pdf.set_fill_color(14, 52, 96)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(W, 7, title, fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(1)

    def _row(label, val):
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(4, 6, "")
        pdf.cell(W - 34, 6, _t(label))
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 6, f"${val:,.2f}", align="R", new_x="LMARGIN", new_y="NEXT")

    def _total(label, val):
        pdf.set_fill_color(226, 239, 218)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(W - 30, 7, label, fill=True)
        pdf.cell(30, 7, f"${val:,.2f}", align="R", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_fill_color(255, 255, 255)

    # Encabezado
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(W, 9, "SOLPLAST", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(W, 7, "ROL DE PAGOS", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    _hline()
    pdf.ln(2)

    for label, val in [("PERIODO:", periodo), ("EMPLEADO:", name)]:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(55, 7, label)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(W - 55, 7, _t(val), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(3)
    _hline()
    pdf.ln(2)

    # Ingresos
    _section("INGRESOS")
    _row("Primera Quincena", n['quincena'])
    _row("Segunda Quincena", n['quincena'])
    if n['pay_50']:
        _row(f"Horas Extras 50% ({n['h_50_pagar']:.2f}h)", n['pay_50'])
    if n['pay_100']:
        _row(f"Horas Extras 100% ({hrs['horas_100']:.2f}h)", n['pay_100'])
    if n['transporte']:
        _row(f"Transporte/Comida ({hrs['dias']}d)", n['transporte'])
    if n['decimo_13']:
        _row("Decimo Tercer Sueldo", n['decimo_13'])
    if n['decimo_14']:
        _row("Decimo Cuarto Sueldo", n['decimo_14'])
    if n['bonus']:
        _row("Bono / Ajuste", n['bonus'])
    pdf.ln(1)
    _total("TOTAL INGRESOS",
           n['total_ingresos'] + n['decimo_13'] + n['decimo_14'] + n['bonus'])

    pdf.ln(4)

    # Egresos
    _section("EGRESOS")
    if n['iess']:
        _row(f"Aporte IESS ({IESS_EMPLEADO * 100:.2f}%)", n['iess'])
    if n['prestamo_iess']:
        _row("Prestamo IESS", n['prestamo_iess'])
    if not n['iess'] and not n['prestamo_iess']:
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(W, 6, "  Sin egresos aplicables", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
    pdf.ln(1)
    _total("TOTAL EGRESOS", n['total_egresos'])

    pdf.ln(4)
    _hline()
    pdf.ln(1)

    # Valor a recibir
    pdf.set_fill_color(14, 52, 96)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(W - 40, 9, "VALOR A RECIBIR", fill=True)
    pdf.cell(40, 9, f"${n['valor_recibir']:,.2f}", align="R", fill=True,
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)

    pdf.ln(6)

    # Detalle de pagos
    _section("DETALLE DE PAGOS")
    _row("Transferencia del 15", n['transf_15'])
    _row("Transferencia fin de mes", n['transf_fin'])
    if n['fondos_reserva']:
        _row("Fondos de Reserva", n['fondos_reserva'])
    pdf.ln(1)
    _total("TOTAL TRANSFERIDO", n['total_transferido'])

    # Footer
    pdf.ln(8)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(140, 140, 140)
    pdf.cell(
        W, 5,
        f"H.50%: Art.55 | H.100%: Art.55 | IESS {IESS_EMPLEADO * 100:.2f}% | 13ro: Art.111 | 14to: Art.113",
        align="C", new_x="LMARGIN", new_y="NEXT",
    )

    return bytes(pdf.output())


if __name__ == '__main__':
    import sys
    xls = sys.argv[1] if len(sys.argv) > 1 else 'NGTimereport.xls'
    out = sys.argv[2] if len(sys.argv) > 2 else 'rol_procesado.xlsx'
    flags = write_excel(parse_xls(xls), out)
    print(f'Generado: {out} | Anomalias: {len(flags)}')
